# Python 3.10.11
# Creado: 31/01/2024
"""Genera un archivo Excel para cambiar las categorías de los eventos."""
import os
import sys

MARX_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__) + "/marx"))
sys.path.append(os.path.dirname(MARX_DIR))


import time
from pathlib import Path

import openpyxl

from marx.model import MarxAdapter
from marx.util import get_most_recent_db


MARX_AT_DESKTOP = "C:/Users/angel/Desktop/marx"


def prepare_excel(adapter: MarxAdapter, dir: str | Path) -> Path:
    """Genera el Excel en la carpeta especificada.

    El excel se compone de 3 hojas: ingresos, gastos y traslados. Cada hoja
    presenta, para cada evento: ID, origen, destino, concepto y la categoría
    actual, más una columna para indicar la nueva.

    Devuelve la ruta del archivo generado.

    """
    output = Path(dir) / f"recat_{int(time.time())}.xlsx"
    wb = openpyxl.Workbook()
    income_sheet = wb.create_sheet("Ingresos")
    expense_sheet = wb.create_sheet("Gastos")
    transfer_sheet = wb.create_sheet("Traslados")
    for type, sheet in zip((1, 0, -1), (income_sheet, transfer_sheet, expense_sheet)):
        sheet.append(("ID", "Origen", "Destino", "Concepto", "Categoría actual", "Nueva categoría"))
        for event in adapter.suite.events.search(type=type):
            sheet.append(
                (
                    str(event.id),
                    event.orig if isinstance(event.orig, str) else event.orig.name,
                    event.dest if isinstance(event.dest, str) else event.dest.name,
                    event.concept,
                    event.category.name,
                    "",
                )
            )
    wb.save(output)
    return output


if __name__ == "__main__":
    source = get_most_recent_db("G:/Mi unidad/MiBilletera Backups")
    adapter = MarxAdapter(source)
    adapter.load()
    excel = prepare_excel(adapter, MARX_AT_DESKTOP)
    print(f"HECHO: {excel}")
