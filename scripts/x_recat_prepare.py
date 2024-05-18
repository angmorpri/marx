# Python 3.10.11
# Creado: 31/01/2024
"""Genera un archivo Excel para cambiar las categorías de los eventos."""
import os
import sys

MARX_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__) + "/marx"))
sys.path.append(os.path.dirname(MARX_DIR))


import time
from pathlib import Path

import openpyxl

from marx.model import MarxMapper
from marx.util import get_most_recent_db


MARX_AT_DESKTOP = "C:/Users/angel/Desktop/marx"


def prepare_excel(adapter: MarxMapper, dir: str | Path) -> Path:
    """Genera el Excel en la carpeta especificada.

    El excel se compone de 3 hojas: ingresos, gastos y traslados. Cada hoja
    presenta, para cada evento: ID, origen, destino, concepto y la categoría
    actual, más una columna para indicar la nueva.

    Devuelve la ruta del archivo generado.

    """
    output = Path(dir) / f"recat_{int(time.time())}.xlsx"
    wb = openpyxl.Workbook()
    income_sheet = wb.create_sheet("Ingresos")
    expense_sheet = wb.create_sheet("Gastos")
    transfer_sheet = wb.create_sheet("Traslados")
    del wb["Sheet"]

    # Ingresos y traslados se representan uno a uno
    for type, sheet in zip((1, 0), (income_sheet, transfer_sheet)):
        sheet.append(("ID", "Origen", "Destino", "Concepto", "Categoría actual", "Nueva categoría"))
        for event in adapter.data.events.search(type=type):
            sheet.append(
                (
                    str(event.id),
                    event.orig if isinstance(event.orig, str) else event.orig.name,
                    event.dest if isinstance(event.dest, str) else event.dest.name,
                    event.concept,
                    event.category.name,
                    "",
                )
            )
    wb.save(output)

    # Gastos se representan agrupados por concepto
    expense_sheet.append(
        ("ID", "Origen", "Destino", "Concepto", "Categoría actual", "Nueva categoría")
    )
    grouped = {}
    for event in adapter.data.events.search(type=-1):
        if event.concept not in grouped:
            grouped[event.concept] = {
                "id": set(),
                "orig": set(),
                "dest": set(),
                "category": set(),
            }
        grouped[event.concept]["id"].add(str(event.id))
        grouped[event.concept]["orig"].add(
            event.orig if isinstance(event.orig, str) else event.orig.name
        )
        grouped[event.concept]["dest"].add(
            event.dest if isinstance(event.dest, str) else event.dest.name
        )
        grouped[event.concept]["category"].add(event.category.name)
    for concept, data in grouped.items():
        expense_sheet.append(
            (
                "; ".join(data["id"]),
                "; ".join(data["orig"]),
                "; ".join(data["dest"]),
                concept,
                "; ".join(data["category"]),
                "",
            )
        )
    wb.save(output)

    return output


if __name__ == "__main__":
    source = get_most_recent_db("G:/Mi unidad/MiBilletera Backups")
    adapter = MarxMapper(source)
    adapter.load()
    excel = prepare_excel(adapter, MARX_AT_DESKTOP)
    print(f"HECHO: {excel}")
