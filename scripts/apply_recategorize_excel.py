# Python 3.10.11
# Creado: 31/01/2024
"""Genera un archivo Excel para cambiar las categorías de los eventos."""
import os
import sys

MARX_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__) + "/marx"))
sys.path.append(os.path.dirname(MARX_DIR))


import time
from pathlib import Path

import openpyxl
from more_itertools import repeat_last

from marx.model import MarxAdapter
from marx.util import get_most_recent_db


MARX_AT_DESKTOP = Path("C:/Users/angel/Desktop/marx")


def apply_excel(adapter: MarxAdapter, dir: str | Path) -> None:
    """Aplica los cambios en el archivo Excel especificado.

    Aplica todo cambio indicado en "Nueva categoría", y también si hay ajustes
    en "Concepto".

    """
    wb = openpyxl.load_workbook(dir)
    income_sheet = wb["Ingresos"]
    # expense_sheet = wb["Gastos"]
    transfer_sheet = wb["Traslados"]
    for sheet in (income_sheet, transfer_sheet):
        for id, *row in sheet.iter_rows(min_row=2, values_only=True):
            if isinstance(id, str):
                id = complex(id) if "j" in id else int(id)
            else:
                id = int(id)
            event = adapter.suite.events[id]
            if not event:
                raise ValueError(f"Evento {id} no encontrado ({row})")
            _, _, concept, _, new_category_id, *_ = row
            if concept != event.concept:
                print(f"[{id}] Concepto cambiado: {event.concept} -> {concept}")
                event.concept = concept
            if new_category_id:
                new_category = adapter.suite.categories[new_category_id]
                if not new_category:
                    raise ValueError(f"[{id}] Categoría no encontrada: {new_category_id}")
                else:
                    print(f"[{id}] Categoría cambiada: {event.category} -> {new_category}")
                    event.category = new_category
        print("\n----------------------------------------------\n")


def expenses_check(adapter: MarxAdapter, dir: str | Path) -> None:
    """Análisis de las nuevas categorías de los gastos en comparación con
    las reales y las usadas.

    """
    wb = openpyxl.load_workbook(dir)
    sheet = wb["Gastos"]
    all_currents = set()
    all_news = set()
    for ids, *row in sheet.iter_rows(min_row=2, values_only=True):
        *_, currents, _, news = row
        all_currents |= set(currents.split("; "))
        all_news |= set(news.split("; "))
    # Categorías teóricas actuales
    tcats = adapter.suite.categories.search(type=-1)
    print("Categorías teóricas actuales: ", len(tcats))
    for category in tcats.sort("code"):
        print(f"  - {category.name}")
    input("\n")
    # Categorías en uso actualmente
    print("Categorías en uso actualmente: ", len(all_currents))
    for category in sorted(all_currents):
        print(f"  - {category}")
    input("\n")
    # Categorías nuevas
    print("Categorías nuevas: ", len(all_news))
    for category in sorted(all_news):
        print(f"  - {category}")
    input("\n")


def apply_expenses(adapter: MarxAdapter, dir: str | Path) -> None:
    """Aplica los cambios y ajustes en los gastos."""
    categories = adapter.suite.categories
    wb = openpyxl.load_workbook(dir)

    # Cambios en los eventos
    sheet = wb["Gastos"]
    event_changes = {}
    for ids, *row in sheet.iter_rows(min_row=2, values_only=True):
        _, dest, concept, current_cat, new_cat, _ = row
        extra_changes = {}
        if dest.startswith("$"):
            extra_changes["dest"] = dest[1:].strip()
        if concept.startswith("$"):
            extra_changes["concept"] = concept[1:].strip()
        new_cats = repeat_last(new_cat.split("; "))
        cat_changes = {old_cat: next(new_cats) for old_cat in current_cat.split("; ")}
        for sid in ids.split("; "):
            id = complex(sid)
            event = adapter.suite.events[id]
            new_cat = cat_changes[event.category.name]
            event_changes[id] = (new_cat, extra_changes)

    # Cambios en las categorías
    sheet = wb["Categorías"]
    changes = {}
    news = []
    for prev_code, new_name in sheet.iter_rows(min_row=2, values_only=True):
        if prev_code == "-":
            news.append(new_name.split(". "))
        else:
            cid = categories[prev_code].id
            changes[cid] = new_name.split(". ")
    for cid, (code, title) in changes.items():
        category = categories[cid]
        category.code = code
        category.title = title
    for code, title in news:
        categories.new(id=-1, name=f"{code}. {title}")
    for category in categories.sort("code"):
        print(category)
    print()

    # Aplicar cambios en los eventos
    for id, (new_cat, extra_changes) in event_changes.items():
        event = adapter.suite.events[id]
        extra_changes["category"] = categories.get(name=new_cat).entity
        print(">>>", extra_changes["category"])
        event.update(**extra_changes)


if __name__ == "__main__":
    source = get_most_recent_db("G:/Mi unidad/MiBilletera Backups")
    print(">>> Usando: ", source)
    time.sleep(1)

    adapter = MarxAdapter(source)
    adapter.load()

    if 0:
        apply_excel(adapter, MARX_AT_DESKTOP / "recat_1706724239.xlsx")
        adapter.suite.categories.search(code="XXX").delete()
        adapter.save()

    if 1:
        # expenses_check(adapter, MARX_AT_DESKTOP / "recat_oficial.xlsx")
        apply_expenses(adapter, MARX_AT_DESKTOP / "recat_oficial.xlsx")
        adapter.save()

    print("HECHO")
