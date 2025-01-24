# Python 3.10.11
# Creado: 23/07/2024
"""Reorganización de categorías de julio de 2024"""
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from marx import Marx
from marx.models import MarxDataStruct

MARX_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__) + "/marx"))
sys.path.append(os.path.dirname(MARX_DIR))


DECEMBER_FILE = r"G:\Mi unidad\MiBilletera Backups\Dic_01_2024_ExpensoDB"

CONCEPT_REPLACEMENTS = {
    "\n": " | ",
    "á": "a",
    "é": "e",
    "í": "i",
    "ó": "o",
    "ú": "u",
}


def prepare_concept(concept: str) -> str:
    """Prepara el concepto para ordenarlo alfabéticamente"""
    concept = concept.lower()
    for old, new in CONCEPT_REPLACEMENTS.items():
        concept = concept.replace(old, new)
    return concept


def organize(data: MarxDataStruct, account: str) -> None:
    """Organiza los datos en `account`, del 1 de enero al 1 de diciembre de
    2024

    Devuelve un diccionario con cuatro entradas:
    - `loss`: Gastos, que es a su vez otro diccionario que agrupa los gastos
        por fecha, categoría y concepto (sólo los gastos, no los traslados)
    - `quota`: Cuotas totales recibidas
    - `extra_profit`: Ingresos fuera de cuotas y traslados, agrupados por
        fecha, categoría y concepto
    - `transfers`: Traslados, divididos en `inbound` y `outbound`, y agrupados
        por fecha, destino u origen, y concepto

    """
    res = {}

    # Gastos
    loss = {}
    for event in data.events.subset(
        lambda x: (
            (datetime(2024, 12, 1) > x.date >= datetime(2023, 12, 1))
            & (x.orig.name == account)
            & (x.flow == x.EXPENSE)
            & (x.status == 1)
        )
    ):
        key = (event.date, event.category.name, prepare_concept(event.concept))
        loss[key] = loss.get(key, 0) + event.amount
    res["loss"] = dict(
        sorted(loss.items(), key=lambda x: (x[1], x[0][0]), reverse=True)
    )

    # Cuotas
    res["quota"] = 0
    for event in data.events.subset(
        lambda x: (
            (datetime(2024, 12, 1) > x.date >= datetime(2023, 12, 1))
            & (x.dest.name == account)
            & (x.category.code == "T11")
            & (x.status == 1)
        )
    ):
        res["quota"] += event.amount

    # Ingresos extra
    profit = {}
    for event in data.events.subset(
        lambda x: (
            (datetime(2024, 12, 1) > x.date >= datetime(2023, 12, 1))
            & (x.dest.name == account)
            & (x.flow == x.INCOME)
            & (x.status == 1)
        )
    ):
        key = (event.date, event.category.name, prepare_concept(event.concept))
        profit[key] = profit.get(key, 0) + event.amount
    res["profit"] = dict(
        sorted(profit.items(), key=lambda x: (x[1], x[0][0]), reverse=True)
    )

    # Traslados
    transfers = {"inbound": {}, "outbound": {}}
    for event in data.events.subset(
        lambda x: (
            (datetime(2024, 12, 1) > x.date >= datetime(2023, 12, 1))
            & ((x.orig.name == account) or (x.dest.name == account))
            & (x.flow == x.TRANSFER)
            & (x.category.code != "T11")
            & (x.status == 1)
        )
    ):
        if event.orig.name == account:
            key = (event.date, event.dest.name, event.category.name)
            transfers["outbound"][key] = (
                transfers["outbound"].get(key, 0) + event.amount
            )
        elif event.dest.name == account:
            key = (event.date, event.orig.name, event.category.name)
            transfers["inbound"][key] = transfers["inbound"].get(key, 0) + event.amount
    res["transfers"] = transfers

    # Return
    res["name"] = account
    return res


def savings(data: MarxDataStruct) -> None:
    """Calcula el ahorro e inversiones totales, por año"""
    # Reserva - Ingresos (incluyendo traslados inbound)
    savings = {}
    for event in data.events.subset(
        lambda x: (
            (x.orig.name == "Reserva" or x.dest.name == "Reserva") and x.status == 1
        )
    ):
        # Ingreso o gasto
        if event.orig.name == "Reserva":
            flow = "loss"
        elif event.dest.name == "Reserva":
            flow = "profit"
        # Categoría
        category = event.category.name
        if event.flow == event.TRANSFER:
            category = f"{event.category.name} @ {event.orig.name}"
        # Añadimos la entrada
        key = (
            f"{event.date.year}-{event.date.month:02d}-01",
            "reserva",
            flow,
            category,
            prepare_concept(event.concept),
        )
        savings[key] = savings.get(key, 0) + event.amount

    # Inversión - Ingresos divididos por categoría
    for event in data.events.subset(
        lambda x: (x.dest.name == "Inversión" and x.status == 1)
    ):
        # Cuenta
        account = prepare_concept(event.concept)
        # Categoría
        category = event.category.name
        if event.flow == event.TRANSFER:
            category = f"{event.category.name} @ {event.orig.name}"
        # Añadimos la entrada
        key = (
            f"{event.date.year}-{event.date.month:02d}-01",
            account,
            "profit",
            category,
            "",
        )
        savings[key] = savings.get(key, 0) + event.amount

    return savings


def create_datasheet(path: Path, currents: dict, savings: dict) -> None:
    """Crea una hoja de cálculo Excel con los datos proporcionados

    Por cada conjunto de datos pasados, creará un resumen en la primera hoja,
    y un detalle en una nueva pestaña.

    """
    wb = openpyxl.Workbook()

    summary_sheet = wb[wb.sheetnames[0]]
    summary_sheet.title = "Resumen"

    for data in currents:
        details_sheet = wb.create_sheet(title=f"Detalle {data['name']}")
        details_sheet.append(
            ["Tipo", "Fecha", "Identificador 1", "Identificador 2", "Cantidad"]
        )

        summary = []
        summary.append((5, 1, "Resultado", 0))
        summary.append((7, 1, "Estado real", 0))
        for item, value in data.items():
            if item == "name":
                summary.append((1, 1, value))

            elif item in ("loss", "profit"):
                # Gastos o ingresos
                title = "Gastos" if item == "loss" else "Ingresos extraordinarios"
                factor = -1 if item == "loss" else 1
                idx = 3 if item == "loss" else 4
                # Detalle y agrupación
                per_category = {}
                for (date, category, concept), amount in value.items():
                    details_sheet.append(
                        [item, date, category, concept, factor * amount]
                    )
                    per_category[category] = per_category.get(category, 0) + amount
                # Resumen
                summary.append((idx, 1, title, sum(per_category.values())))
                for i, (category, amount) in enumerate(
                    sorted(per_category.items(), key=lambda x: x[1], reverse=True)
                ):
                    summary.append((idx, i + 2, category, amount))

            elif item == "quota":
                summary.append((2, 1, "Cuotas", value))

            elif item == "transfers":
                per_target = {}
                for direction, transfers in value.items():
                    factor = -1 if direction == "outbound" else 1
                    for (date, account, category), amount in transfers.items():
                        details_sheet.append(
                            [f"transfer.{direction}", date, account, category, amount]
                        )
                        target = (
                            f"A {account}"
                            if direction == "outbound"
                            else f"Desde {account}"
                        )
                        per_target[target] = per_target.get(target, 0) + factor * amount
                # Resumen
                summary.append((6, 1, "Traslados", sum(per_target.values())))
                for i, (target, amount) in enumerate(
                    sorted(per_target.items(), key=lambda x: x[1], reverse=True)
                ):
                    summary.append((6, i + 2, target, amount))

        # Resumen
        for row in sorted(summary, key=lambda x: (x[0], x[1])):
            summary_sheet.append(row[2:])
        summary_sheet.append([])

    # Ahorros e inversiones
    savings_sheet = wb.create_sheet(title="Detalle de inversiones")
    savings_sheet.append(["Año", "Cuenta", "Tipo", "Categoría", "Concepto", "Cantidad"])
    for key, value in sorted(savings.items(), key=lambda x: x[1], reverse=True):
        savings_sheet.append([key[0], key[1], key[2], key[3], key[4], value])

    # Guardar
    wb.save(path)
    wb.close()


if __name__ == "__main__":
    marx = Marx()
    marx.load(DECEMBER_FILE)
    data = marx.data

    # Básicos y personales
    bres = organize(data, "Básicos")
    pres = organize(data, "Personales")
    ares = organize(data, "Reserva")

    # Ahorros e inversiones
    rres = savings(data)
    for key, value in sorted(rres.items(), key=lambda x: x[1], reverse=True):
        print(key, "      ", value)

    # Crear hoja de cálculo
    create_datasheet(
        r"C:\Users\angel\Playground\finances2025.xlsx", (bres, pres, ares), rres
    )
    print("DONE")
