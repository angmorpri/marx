# 3.10.11
# Creado: 07/03/2024
"""Estilos para celdas y tablas de Excel."""

from __future__ import annotations

from copy import deepcopy
from pathlib import Path
from typing import Any, Literal

import openpyxl
from openpyxl.cell.cell import Cell as OpenpyxlCell
from openpyxl.styles import Alignment as OpenpyxlAlignment
from openpyxl.styles import Border as OpenpyxlBorder
from openpyxl.styles import Color as OpenpyxlColor
from openpyxl.styles import Font as OpenpyxlFont
from openpyxl.styles import PatternFill as OpenpyxlFill
from openpyxl.styles import Side as OpenpyxlSide


RGB = str
ExcelTheme = tuple[int, int]

WHITE = "#FFFFFF"
BLACK = "#000000"
GRAY = (1, 0.85)


class Color:
    """Representa un color en los formatos válidos de Excel.

    Los formatos válidos son dos:
        - RGB: cadena con forma "#RRGGBB".
        - Tema: tupla formada por (índice de tema, intensidad de tema).

    Mediante el método 'get()', se obtiene un objeto 'openpyxl.styles.Color'
    que se puede usar para asignar color a celdas y tablas.

    El constructor admite cualquiera de los dos tipos válidos.

    """

    def __init__(self, color: RGB | ExcelTheme | Color) -> None:
        self.rgb = None
        self.theme = None
        self.tint = None
        if isinstance(color, str):
            self.rgb = color
        elif isinstance(color, tuple):
            self.theme, self.tint = color
        elif isinstance(color, Color):
            self.rgb = color.rgb
            self.theme = color.theme
            self.tint = color.tint
        else:
            raise ValueError(f"El color '{color}' no es válido.")

    @classmethod
    def from_color(
        cls, color: OpenpyxlColor | None, default: RGB | ExcelTheme | Color | None = None
    ) -> Color:
        """Copia un objeto 'openpyxl.styles.Color'."""
        if color is None:
            return cls(default)
        elif rgb := color.__dict__.get("rgb", False):
            return cls(f"#{rgb[2:]}")
        elif theme := color.__dict__.get("theme", False):
            tint = color.__dict__.get("tint", 0.0)
            return cls((theme, tint))
        else:
            return cls(default)

    def get(self) -> OpenpyxlColor:
        """Devuelve un objeto 'openpyxl.styles.Color'."""
        if self.rgb:
            return OpenpyxlColor(rgb=f"FF{self.rgb[1:]}")
        return OpenpyxlColor(theme=self.theme, tint=self.tint)

    def __repr__(self) -> str:
        if self.rgb:
            return f"Color({self.rgb})"
        return f"Color({self.theme}, {self.tint:.2f})"

    def __str__(self) -> str:
        if self.rgb:
            return self.rgb
        return f"<{self.theme}, {self.tint:.2f}>"


class HandlesColor:
    """Clase base para objetos que manejan color.

    Proporciona el atributo protegido '_color', y la propiedad 'color', que
    garantiza que el valor asignado sea un objeto 'Color', convirtiéndolo si
    es necesario.

    """

    def __init__(self, color: Color | RGB | ExcelTheme) -> None:
        self._color = Color(color)

    @property
    def color(self) -> Color:
        """Devuelve el color."""
        return self._color

    @color.setter
    def color(self, value: Color | RGB | ExcelTheme) -> None:
        """Establece el color."""
        self._color = Color(value)


class Border(HandlesColor):
    """Representa uno de los cuatro bordes de una celda en Excel.

    Cada borde se compone de un color ('color') y un indicador de si se debe
    ocultar o no ('hide'). El color puede ser un objeto 'Color', una cadena
    para indicar el color RGB, o un tema de Excel.

    Se proporciona el método 'get()' para obtener un objeto
    'openpyxl.styles.Side'.

    El constructor debe recibir el color en cualquier de sus formatos válidos
    y un valor booleano para 'hide' (por defecto, False).

    """

    def __init__(self, color: Color | RGB | ExcelTheme = WHITE, hide: bool = False) -> None:
        super().__init__(color)
        self.hide = hide

    def get(self) -> OpenpyxlSide:
        """Devuelve un objeto 'openpyxl.styles.Side'."""
        if self.hide:
            return OpenpyxlSide(color=None, border_style=None)
        return OpenpyxlSide(color=self.color.get(), border_style="thin")

    def __str__(self) -> str:
        if self.hide:
            return f"HiddenBorder()"
        return f"Border({self.color!s})"


class Borders:
    """Representa y almacena los cuatro bordes de una celda en Excel.

    Se compone de cuatro objetos 'Border' que representan los bordes superior
    ('top'), inferior ('bottom'), izquierdo ('left') y derecho ('right') de la
    celda.

    Se proporciona el método 'get()' para obtener un objeto
    'openpyxl.styles.Border'.

    El constructor admite, para cada borde:
        - Un objeto 'Border'.
        - Un color en cualquiera de sus formatos válidos.
    Además, mediante el argumento clave 'hide', se puede especificar qué bordes
    ocultar. Éste debe ser una cadena formada por las letras 't', 'b', 'l' o
    'r', en función de si se quiere ocultar el borde superior, inferior,
    izquierdo o derecho, respectivamente.

    También se puede crear a partir de una celda con formato Openpyxl mediante
    el método de clase 'from_cell'.

    """

    def __init__(
        self,
        top: Border | Color | RGB | ExcelTheme = GRAY,
        bottom: Border | Color | RGB | ExcelTheme = GRAY,
        left: Border | Color | RGB | ExcelTheme = GRAY,
        right: Border | Color | RGB | ExcelTheme = GRAY,
        *,
        hide: str = "",
    ) -> None:
        self.top = top if isinstance(top, Border) else Border(top)
        self.bottom = bottom if isinstance(bottom, Border) else Border(bottom)
        self.left = left if isinstance(left, Border) else Border(left)
        self.right = right if isinstance(right, Border) else Border(right)
        if hide:
            self.top.hide = "t" in hide
            self.bottom.hide = "b" in hide
            self.left.hide = "l" in hide
            self.right.hide = "r" in hide

    @classmethod
    def from_cell(cls, cell: OpenpyxlCell) -> Borders:
        """Copia los bordes de una celda de Openpyxl."""
        return cls(
            top=Border(
                Color.from_color(cell.border.top.color, GRAY),
                cell.border.top.style is None,
            ),
            bottom=Border(
                Color.from_color(cell.border.bottom.color, GRAY),
                cell.border.bottom.style is None,
            ),
            left=Border(
                Color.from_color(cell.border.left.color, GRAY),
                cell.border.left.style is None,
            ),
            right=Border(
                Color.from_color(cell.border.right.color, GRAY),
                cell.border.right.style is None,
            ),
        )

    def get(self) -> OpenpyxlBorder:
        """Devuelve un objeto 'openpyxl.styles.Border'."""
        return OpenpyxlBorder(
            top=self.top.get(),
            bottom=self.bottom.get(),
            left=self.left.get(),
            right=self.right.get(),
        )

    def __str__(self) -> str:
        code = []
        for side in ("top", "bottom", "left", "right"):
            if not getattr(self, side).hide:
                code.append(f"{side[0]}={getattr(self, side).color!s}")
        return f"Borders({', '.join(code)})"


class TextStyle(HandlesColor):
    """Representa un estilo de texto en Excel.

    Los estilos de texto adminte:
        - texto en negrita, mediante el atributo 'bold'.
        - texto en cursiva, mediante el atributo 'italic'.
        - color de texto, mediante el atributo 'color'. Por defecto, negro.
        - alineamiento horizontal, mediante el atributo 'align', que acepta
            los valores "left", "center" y "right". Por defecto, "left".
        - tabulación, mediante 'indent'. Por defecto, 0.
        - formato numérico, si es necesario. Por defecto, no tendrá.

    El constructor admite cada uno de los atributos listados como argumento.
    Además, en el caso del color, admite cualquier forma válida de color.

    También se puede crear a partir de una celda con formato Openpyxl mediante
    el método de clase 'from_cell'.

    """

    def __init__(
        self,
        bold: bool = False,
        italic: bool = False,
        color: Color | RGB | ExcelTheme = BLACK,
        align: Literal["left", "center", "right"] = "left",
        indent: int = 0,
        format: str = "",
    ) -> None:
        super().__init__(color)
        self.bold = bold
        self.italic = italic
        self.align = align
        self.indent = indent
        self.format = format

    @classmethod
    def from_cell(cls, cell: OpenpyxlCell) -> TextStyle:
        """Copia el estilo de texto de una celda de Openpyxl."""
        return cls(
            bold=cell.font.bold or False,
            italic=cell.font.italic or False,
            color=Color.from_color(cell.font.color, BLACK),
            align=cell.alignment.horizontal or "left",
            indent=cell.alignment.indent or 0,
            format=cell.number_format or "",
        )

    def __str__(self) -> str:
        code = []
        if self.bold:
            code.append("b")
        if self.italic:
            code.append("i")
        code.append(self.align[0].lower())
        code.append(str(self.indent))
        if self.format:
            code.append("#")
        return f"TextStyle({self.color!s}, {''.join(code)})"


class CellStyle(HandlesColor):
    """Representa un estilo de celda en Excel.

    Se compone de:
        - un estilo de texto 'TextStyle', mediante el atributo 'text'.
        - un color de fondo, mediante el atributo 'color'.
        - un borde 'Borders', mediante el atributo 'borders'.

    También, opcionalmente, pueden tener un nombre identificador 'name'.

    Se proporciona el método 'apply', que recibe un objeto 'openpyxl.cell.Cell'
    y aplica el estilo a la celda.

    El constructor admite los siguientes argumentos:
        - 'text': un objeto 'TextStyle', o un diccionario con los atributos
            válidos para el constructor de 'TextStyle'.
        - 'color': un objeto 'Color', o un color en cualquiera de sus formatos
            válidos.
        - 'borders': un objeto 'Borders', o una tupla (color, hide), donde
            'color' será el color para todos los bordes, y 'hide' una cadena
            indicando qué bordes ocultar, siendo 't', 'b', 'l' y 'r' las
            letras válidas.

    También se puede crear a partir de una celda con formato Openpyxl mediante
    el método de clase 'from_cell'.

    """

    def __init__(
        self,
        text: TextStyle | dict[str, Any] = {},
        color: Color | RGB | ExcelTheme = WHITE,
        borders: Borders | tuple[Color | RGB | ExcelTheme, str] = (GRAY, ""),
        *,
        name: str | None = None,
    ) -> None:
        super().__init__(color)
        self.text = text if isinstance(text, TextStyle) else TextStyle(**text)
        if isinstance(borders, Borders):
            self.borders = borders
        else:
            c, hide = borders
            self.borders = Borders(c, c, c, c, hide=hide)
        self.name = name

    @classmethod
    def from_cell(cls, cell: OpenpyxlCell, *args, **kwargs) -> CellStyle:
        """Copia el estilo de celda a partir de una celda de Openpyxl."""
        return cls(
            text=TextStyle.from_cell(cell),
            color=Color.from_color(cell.fill.fgColor, WHITE),
            borders=Borders.from_cell(cell),
            *args,
            **kwargs,
        )

    def apply(self, cell: OpenpyxlCell) -> None:
        """Aplica el estilo a la celda."""
        cell.font = OpenpyxlFont(
            bold=self.text.bold,
            italic=self.text.italic,
            color=self.text.color.get(),
        )
        cell.alignment = OpenpyxlAlignment(
            horizontal=self.text.align,
            vertical="center",
            indent=self.text.indent,
        )
        if self.text.format:
            cell.number_format = self.text.format
        cell.fill = OpenpyxlFill(
            fill_type="solid",
            start_color=self.color.get(),
        )
        cell.border = self.borders.get()

    def __str__(self) -> str:
        if self.name:
            return f"CellStyle({self.name!r}, {self.color!s}, text={self.text!s}, borders={self.borders!s})"
        return f"CellStyle({self.color!s}, text={self.text!s}, borders={self.borders!s})"


class StylesCatalog:
    """Catálogo de estilos de celda.

    Permite almacenar estilos de celda tipo 'CellStyle' bajo un nombre único,
    que será accesible mediante [] o mediante atributos.

    Los estilos pueden agruparse en temas y subtemas para facilitar su
    organización. Para definir un nuevo estilo, se proporciona el método
    'define()', que recibe el nombre del estilo, la ruta de su tema y subtemas,
    y el estilo en sí. Si un estilo con el mismo nombre ya existe, se
    sobrescribirá.

    Alternativamente, un conjunto de estilos y temas pueden cargarse desde un
    archivo Excel mediante el método 'load()'. Se recorrerá cada celda editada
    copiando el estilo de ésta, exceptuando la primera fila y la primera
    columna, que indicarán "<tema>[-<subtema>]" y nombre del estilo
    respectivamente.

    El constructor puede recibir una ruta a un archivo Excel para cargar los
    estilos iniciales.

    """

    def __init__(self, path: str | Path | None = None) -> None:
        self._styles = {}
        if path:
            self.load(path)

    def load(self, path: str | Path) -> None:
        """Carga los estilos desde un archivo Excel."""
        wb = openpyxl.load_workbook(Path(path))
        sheet = wb.active
        themes_by_column = {
            col: sheet.cell(row=1, column=col).value for col in range(2, sheet.max_column + 1)
        }
        names_by_row = {
            row: sheet.cell(row=row, column=1).value for row in range(2, sheet.max_row + 1)
        }
        for row in sheet.iter_rows(min_row=2, min_col=2):
            for cell in row:
                theme = themes_by_column[cell.column]
                name = names_by_row[cell.row]
                style = CellStyle.from_cell(cell, name=f"{theme}-{name}")
                self.define(name, style, theme.split("-"))

    def define(self, name: str, style: CellStyle, themes: str | tuple[str] = "") -> None:
        """Define un nuevo estilo de celda.

        Si se proporcionan temas y subtemas, estos se crearán si no existen.

        """
        if isinstance(themes, str):
            themes = (themes,)
        catalog = self
        for theme in (t for t in themes if t):
            if theme not in catalog:
                catalog._styles[theme] = StylesCatalog()
            catalog = catalog._styles[theme]
        catalog._styles[name] = style

    def _get(self, name: str) -> CellStyle | StylesCatalog:
        """Devuelve el estilo de celda."""
        if name not in self._styles:
            raise KeyError(f"El estilo o tema '{name}' no está definido.")
        if isinstance(self._styles[name], StylesCatalog):
            return self._styles[name]
        return deepcopy(self._styles[name])

    def __getitem__(self, name: str) -> CellStyle | StylesCatalog:
        """Devuelve el estilo de celda."""
        return self._get(name)

    def __getattr__(self, name: str) -> CellStyle | StylesCatalog:
        """Devuelve el estilo de celda."""
        return self._get(name)

    def __contains__(self, name: str) -> bool:
        """Devuelve si el estilo o tema está definido."""
        return name in self._styles

    def _str(self, text: list[str], level: int = 0) -> str:
        """Devuelve una línea de texto."""
        for name, style in self._styles.items():
            if isinstance(style, StylesCatalog):
                text.append(f"{'  ' * level}{name}:")
                style._str(text, level + 1)
            else:
                text.append(f"{'  ' * level}{name}: {style!s}")

    def __str__(self) -> str:
        """Devuelve una representación en forma de texto."""
        text = []
        self._str(text)
        return "\n".join(text)


if __name__ == "__main__":
    path = Path(__file__).parents[3] / "config" / "styles.xlsx"
    styles = StylesCatalog(path)
    print(styles)
