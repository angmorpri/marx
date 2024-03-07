# 3.10.11
# Creado: 03/03/2024
"""Módulo de utilidades y herramientas para manejar hojas de cálculo de Excel
más cómodamente.

Por detrás usa la librería openpyxl para manejar los archivos de Excel. El
objetivo es que el manejo de una hoja de cálculo sea más fiel al que se haría
con una hoja de cálculo de Excel real.

"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell as OpenpyxlCell
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlSheet

from marx.reporting.excel.styles import CellStyle


CellIDLike = str | tuple[int, int]


# Hojas


class Sheet:
    """Clase que representa una página de una hoja de cálculo Excel."""

    def __init__(self, base: Excel, sheet: OpenpyxlSheet) -> None:
        self._base = base
        self._sheet = sheet
        self.cells = {}

    @property
    def raw(self) -> OpenpyxlSheet:
        """Devuelve la página de la hoja de cálculo como un objeto openpyxl."""
        return self._sheet

    @property
    def title(self) -> str:
        """Devuelve el nombre de la página."""
        return self._sheet.title

    @title.setter
    def title(self, new_title: str) -> None:
        """Cambia el nombre de la página."""
        self._sheet.title = new_title

    @property
    def index(self) -> int:
        """Devuelve el índice de la página en la hoja de cálculo."""
        return self._base._wb.index(self._sheet)

    def select(self) -> None:
        """Selecciona la página actual."""
        self._base.current_sheet = self

    def rename(self, new_title: str) -> None:
        """Cambia el nombre de la página actual."""
        self._sheet.title = new_title

    def clean(self) -> None:
        """Elimina todas las celdas de la página actual."""
        self._sheet.delete_rows(1, self.raw.max_row)

    def delete(self) -> None:
        """Elimina la página actual."""
        self._base._wb.remove(self._sheet)

    def set_column_width(self, column: str | int, width: int) -> None:
        """Establece el ancho de una columna."""
        if isinstance(column, int):
            column = chr(column + 64)
        self._sheet.column_dimensions[column.upper()].width = width

    def __getitem__(self, key: CellIDLike | CellID) -> Cell:
        id = CellID(key)
        if id not in self.cells:
            self.cells[id] = Cell(id, self.raw[id.as_str()])
        return self.cells[id]

    def __str__(self):
        return f"Sheet({self.title})"


class Sheets:
    """Clase que representa al conjunto de páginas de una hoja de cálculo."""

    def __init__(self, base: Excel) -> None:
        self._base = base

    def new(self, title: str, position: int | None = None, *, select: bool = False) -> Sheet:
        """Crea una nueva página en la hoja de cálculo.

        Si se indica 'position', la nueva página se insertará en esa posición
        (empezando por 0). Si no se indica, la nueva página se insertará al
        final.

        Si 'select' es True, la nueva página se seleccionará tras su creación.

        Devuelve la nueva página.

        """
        self._base._wb.create_sheet(title=title, index=position)
        sheet = Sheet(self._base, self._base._wb[title])
        if select:
            sheet.select()
        return sheet

    def __getitem__(self, key: int | str) -> Sheet:
        if isinstance(key, str):
            return Sheet(self._base, self._base._wb[key])
        elif isinstance(key, int):
            return Sheet(self._base, self._base._wb[self._base._wb.sheetnames[key]])
        else:
            raise ValueError("El identificador de la página debe ser un entero o una cadena.")

    def __len__(self) -> int:
        return len(self._base._wb.sheetnames)

    def __iter__(self):
        return iter(Sheet(self._base, self._base._wb[sheet]) for sheet in self._base._wb.sheetnames)

    def __str__(self):
        return f"Sheets({len(self)} páginas)"


# Celdas


class CellID:
    """Clase que representa un identificador de una celda arbitraria en una
    hoja de cálculo Excel.

    Los identificadores de celda tienen dos formas: texto y tupla. Por ejemplo,
    la celda A1 puede ser representada como "A1" o (1, 1). Esta clase permite
    convertir entre ambas formas y también permite operar con ellas.

    El constructor de la clase acepta un identificador de celda en cualquiera
    de las dos formas.

    """

    def __init__(self, cell_id: CellIDLike | CellID) -> None:
        if isinstance(cell_id, str):
            cell_id = cell_id.upper()
            self._cell_id = cell_id
            self._cell_tuple = (ord(cell_id[0]) - 64, int(cell_id[1:]))
        elif isinstance(cell_id, tuple):
            self._cell_id = f"{chr(cell_id[0] + 64)}{cell_id[1]}"
            self._cell_tuple = cell_id
        elif isinstance(cell_id, CellID):
            self._cell_id = cell_id._cell_id
            self._cell_tuple = cell_id._cell_tuple
        else:
            raise ValueError("El identificador de celda debe ser una cadena o una tupla.")

    def as_str(self) -> str:
        """Devuelve el identificador de celda como una cadena."""
        return self._cell_id

    def as_id(self) -> str:
        """Devuelve el identificador de celda como una cadena."""
        return self._cell_id

    def as_tuple(self) -> tuple[int, int]:
        """Devuelve el identificador de celda como una tupla."""
        return self._cell_tuple

    def as_coords(self) -> tuple[int, int]:
        """Devuelve las coordenadas de la celda como una tupla."""
        return self._cell_tuple

    def __eq__(self, other: CellID | CellIDLike) -> bool:
        return self._cell_id == CellID(other)._cell_id

    def __lt__(self, other: CellID | CellIDLike) -> bool:
        return self._cell_tuple < CellID(other)._cell_tuple

    def __hash__(self) -> int:
        return hash(self._cell_id)

    def __str__(self):
        return self._cell_id


class Cell:
    """Clase que representa una celda en una hoja de cálculo Excel."""

    def __init__(self, id: CellID, cell: OpenpyxlCell, style: CellStyle | None = None) -> None:
        self.id = id
        self._cell = cell
        self.style = style or CellStyle()

    @property
    def raw(self) -> OpenpyxlCell:
        """Devuelve la celda como un objeto openpyxl."""
        return self._cell

    @property
    def value(self) -> Any:
        """Devuelve el valor de la celda."""
        return self._cell.value

    @value.setter
    def value(self, value: Any) -> None:
        """Modifica el valor de la celda."""
        self._cell.value = value

    def __str__(self):
        return f"Cell({self.id})"


class CellPointer:
    """Clase que representa un puntero a una celda en una hoja de cálculo
    Excel.

    Permite moverse por toda la hoja mediante los métodos 'up', 'down', 'left',
    'right' y 'goto'. Para acceder a la celda apuntada, se proporciona el
    atributo 'cell', de la clase 'Cell'.

    """

    def __init__(self, sheet: Sheet, start: CellID) -> None:
        self._sheet = sheet
        self._current = start
        self._sheet[start]  # Asegura que la celda esté en el diccionario

    @property
    def cell(self) -> Cell:
        """Devuelve la celda apuntada por el puntero."""
        return self._sheet.cells[self._current]

    @property
    def value(self) -> Any:
        """Devuelve el valor de la celda apuntada por el puntero."""
        return self.cell.value

    @value.setter
    def value(self, value: Any) -> None:
        """Modifica el valor de la celda apuntada por el puntero."""
        self.cell.value = value

    @property
    def row(self) -> int:
        """Devuelve el número de fila de la celda apuntada por el puntero."""
        return self._current.as_tuple()[1]

    @property
    def column(self) -> int:
        """Devuelve el número de columna de la celda apuntada por el puntero."""
        return self._current.as_tuple()[0]

    # Movimientos
    def up(self, steps: int = 1) -> CellPointer:
        """Mueve el puntero hacia arriba."""
        new_id = CellID((self._current.as_tuple()[0], self._current.as_tuple()[1] - steps))
        return self.goto(new_id)

    def down(self, steps: int = 1) -> CellPointer:
        """Mueve el puntero hacia abajo."""
        new_id = CellID((self._current.as_tuple()[0], self._current.as_tuple()[1] + steps))
        return self.goto(new_id)

    def left(self, steps: int = 1) -> CellPointer:
        """Mueve el puntero hacia la izquierda."""
        new_id = CellID((self._current.as_tuple()[0] - steps, self._current.as_tuple()[1]))
        return self.goto(new_id)

    def right(self, steps: int = 1) -> CellPointer:
        """Mueve el puntero hacia la derecha."""
        new_id = CellID((self._current.as_tuple()[0] + steps, self._current.as_tuple()[1]))
        return self.goto(new_id)

    def goto(self, cell: CellIDLike | CellID) -> CellPointer:
        """Mueve el puntero a una celda arbitraria."""
        id = CellID(cell)
        self._sheet[id]  # Asegura que la celda esté en el diccionario
        self._current = id
        return self

    # Otros
    def copy(self) -> CellPointer:
        """Crea una copia del puntero."""
        return CellPointer(self._sheet, self._current)

    def __str__(self):
        return f"CellPointer({self._current})"


# Principal


class Excel:
    """Clase que representa un archivo de hoja de cálculo Excel.

    Esencialmente sirve para crear o abrir un archivo de hoja de cálculo Excel,
    y para crear punteros a celdas en una hoja de cálculo. También permite
    ejecutar operaciones de selección, creación y modificación de páginas; y
    también enmascara operaciones propias de openpyxl.

    El constructor requiere la ruta a un archivo de hoja de cálculo Excel. Si
    el archivo no existe, se creará uno nuevo. Si el archivo existe, se abrirá
    el archivo existente. Alternativamente, se pueden usar los métodos de clase
    'new' (que crea un archivo nuevo o sobrescribe uno existente) o 'open' (que
    abre un archivo existente o lanza una excepción si no existe).

    """

    def __init__(self, path: Path | str) -> None:
        # Carga el archivo de hoja de cálculo Excel
        self._path = Path(path)
        if self._path.exists():
            self._wb = openpyxl.load_workbook(self._path)
        else:
            self._wb = openpyxl.Workbook()
            self._wb.save(self._path)

        # Inicializa los atributos necesarios
        self.sheets = Sheets(self)
        self.current_sheet = self.sheets[0]

    @classmethod
    def new(cls, path: Path | str) -> Excel:
        """Crea un archivo nuevo de hoja de cálculo Excel."""
        path = Path(path)
        if path.exists():
            path.unlink()
        return cls(path)

    @classmethod
    def open(cls, path: Path | str) -> Excel:
        """Abre un archivo existente de hoja de cálculo Excel."""
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"No existe el archivo '{path}'.")
        return cls(path)

    # Control general
    def save(self) -> None:
        """Guarda el archivo de hoja de cálculo Excel."""
        self._wb.save(self._path)

    def close(self) -> None:
        """Cierra el archivo de hoja de cálculo Excel."""
        self._wb.close()

    # Métodos de celdas
    def pointer(self, at: CellIDLike | CellID) -> CellPointer:
        """Crea un puntero a una celda en la página actual."""
        return CellPointer(self.current_sheet, CellID(at))

    # Estilos
    def stylize(self) -> None:
        """Aplica los estilos asignados a cada una de las celdas creadas."""
        for sheet in self.sheets:
            print(">>>", sheet.title, len(sheet.cells))
            for cell in sheet.cells.values():
                print(">>>>>>", sheet.title, cell.id.as_str(), cell.style)
                cell.style.apply(cell)

    # Utilidades y herramientas
    def compose_formula(self, formula: str, cells: list[CellIDLike | CellID]) -> str:
        """Compone una fórmula a partir de una lista de identificadores de
        celdas.

        La fórmula es la especificada en 'formula' y los identificadores de
        celdas son los especificados en 'cells'.

        La fórmula resultante es capaz de convertir la lista de identificadores
        en rangos de celdas allí donde sea posible.

        """
        cells = list(sorted(CellID(cell) for cell in cells))
        cell_ranges = []
        prev_cell = None
        for cell in cells:
            if prev_cell is None:
                prev_cell = cell
                cell_ranges.append([cell, None])
                continue
            x, y = cell.as_tuple()
            px, py = prev_cell.as_tuple()
            if abs(x - px) == 1 and y == py:
                cell_ranges[-1][1] = cell
            elif abs(y - py) == 1 and x == px:
                cell_ranges[-1][1] = cell
            else:
                cell_ranges.append([cell, None])
            prev_cell = cell
        formatted_cells = []
        for start, end in cell_ranges:
            if end is None:
                formatted_cells.append(start.as_str())
            else:
                formatted_cells.append(f"{start.as_str()}:{end.as_str()}")
        return f"={formula}({','.join(formatted_cells)})"

    # Otros
    def __str__(self):
        return f"Excel({self._path})"
