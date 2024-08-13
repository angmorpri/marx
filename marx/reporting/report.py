# Python 3.10.11
# Creado: 13/08/2024
"""Clase base y herramientas básicas para la preparación de informes

Presenta la clase 'Report', de la que debe heredar cualquier otra clase que
quiera ser usada como generador de informes.

"""

from abc import ABC, abstractmethod
from datetime import datetime
from pathlib import Path

import openpyxl

from marx.models import MarxDataStruct


class Report(ABC):
    """Clase base para la generación de informes

    Un informe es un documento que recibe todos los datos contables cargados
    de una base de datos (en formato 'MarxDataStruct') y, junto con una serie
    de fechas, debe generar una hoja de cálculo Excel exponiendo la información
    deseada.

    Todo informe debe poser un nombre identificador ('name'), y, opcionalmente,
    un título ('title'), una descripción ('description') y un nombre especial
    para la hoja Excel en la que se muestren los datos ('sheet_name'). Si no
    se indican estos últimos, se infieren a partir del nombre.

    El constructor recibirá únicamente la estructura de datos contables que
    esté en uso en el sistema ('data').

    Las clases hijas deberán implementar el método 'build', que recibirá una
    lista de fechas y generará el informe, en formato Excel, en la hoja
    apuntada por el atributo 'sheet', que, por defecto, tendrá formato de
    OpenPyXL. Este atributo 'sheet' es creado por el método 'prepare', al que
    se le pasa la ruta del archivo Excel de salida y un identificador de la
    hoja en la que debe escribirse. Por defecto, esta clase garantizará que
    existe el archivo, o lo creará si no es así; y también creará la hoja
    requerida si no existe, o la limpiará si ya está creada. En cualquier caso,
    si se quiere seguir otro mecanismo o utilizar otra librería para controlar
    la hoja Excel, las clases hijas pueden sobrescribir este método también.

    Finalmente, el método 'save' se encargará de guardar el archivo Excel con
    el informe generado. Si se ha modificado 'prepare' o la forma de gestionar
    la Excel, es posible que también haya que modificar este método.

    """

    def __init__(self, data: MarxDataStruct) -> None:
        self.data = data
        if not hasattr(self, "name"):
            raise AttributeError("[MarxReport] El informe debe tener un nombre")
        if not hasattr(self, "title"):
            self.title = self.name
        if not hasattr(self, "description"):
            self.description = ""
        if not hasattr(self, "sheet_name"):
            self.sheet_name = self.name.lower().replace(" ", "_")

    def prepare(self, path: Path, sheet_id: int | str | None = None) -> None:
        """Prepara la hoja Excel donde se escribirá el informe

        Si el archivo no existe, lo crea. Si 'sheet_id' es un entero, buscará
        la hoja con dicho índice, o el máximo si es mayor. Si es un nombre,
        buscará la hoja con dicho nombre, o la creará si no existe. Si es None,
        creará una nueva hoja, poniendo como nombre aquél que indique el
        atributo 'sheet_name'.

        La hoja generada, en formato OpenPyXL, se guardará en el atributo
        'sheet'. Todo su contenido será eliminado si ya existía.

        """
        if not path.exists():
            wb = openpyxl.Workbook()
            wb.save(path)
            wb.close()
        self.excel_path = path
        # ubicar la hoja
        wb = openpyxl.load_workbook(path)
        sheet_id = self.sheet_name if sheet_id is None else sheet_id
        if isinstance(sheet_id, str):
            if sheet_id not in wb.sheetnames:
                wb.create_sheet(title=sheet_id)
            self.sheet = wb[sheet_id]
        elif isinstance(sheet_id, int):
            self.sheet = wb.worksheets[max(sheet_id, len(wb.sheetnames) - 1)]
        # limpiar la hoja
        self.sheet.delete_rows(1, self.sheet.max_row)
        self._wb = wb

    def save(self) -> Path:
        """Guarda el archivo Excel con el informe generado"""
        self._wb.save(self.excel_path)
        self._wb.close()
        return self.excel_path

    @abstractmethod
    def build(self, dates: list[datetime]) -> None:
        """Genera el informe

        Debe ser implementado por las clases hijas.

        """
        pass

    def __str__(self) -> str:
        return f"Report({self.name}, {self.title!r}, {self.description!r}, {self.sheet_name!r})"
