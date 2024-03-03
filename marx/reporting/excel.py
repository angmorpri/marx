# 3.10.11
# Creado: 03/03/2024
"""Módulo de utilidades y herramientas para manejar hojas de cálculo de Excel
más cómodamente.

Por detrás usa la librería openpyxl para manejar los archivos de Excel. El
objetivo es que el manejo de una hoja de cálculo sea más fiel al que se haría
con una hoja de cálculo de Excel real.

"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell as OpenpyxlCell
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlSheet


CellIDLike = str | tuple[int, int]


class CellID:
    """Clase que representa un identificador de una celda arbitraria en una
    hoja de cálculo Excel.

    Los identificadores de celda tienen dos formas: texto y tupla. Por ejemplo,
    la celda A1 puede ser representada como "A1" o (1, 1). Esta clase permite
    convertir entre ambas formas y también permite operar con ellas.

    El constructor de la clase acepta un identificador de celda en cualquiera
    de las dos formas.

    """

    def __init__(self, cell_id: CellIDLike | CellID) -> None:
        if isinstance(cell_id, str):
            cell_id = cell_id.upper()
            self._cell_id = cell_id
            self._cell_tuple = (ord(cell_id[0]) - 64, int(cell_id[1:]))
        elif isinstance(cell_id, tuple):
            self._cell_id = f"{chr(cell_id[0] + 64)}{cell_id[1]}"
            self._cell_tuple = cell_id
        elif isinstance(cell_id, CellID):
            self._cell_id = cell_id._cell_id
            self._cell_tuple = cell_id._cell_tuple
        else:
            raise ValueError("El identificador de celda debe ser una cadena o una tupla.")

    def as_str(self) -> str:
        """Devuelve el identificador de celda como una cadena."""
        return self._cell_id

    def as_id(self) -> str:
        """Devuelve el identificador de celda como una cadena."""
        return self._cell_id

    def as_tuple(self) -> tuple[int, int]:
        """Devuelve el identificador de celda como una tupla."""
        return self._cell_tuple

    def as_coords(self) -> tuple[int, int]:
        """Devuelve las coordenadas de la celda como una tupla."""
        return self._cell_tuple

    def __eq__(self, other: CellID | CellIDLike) -> bool:
        return self._cell_id == CellID(other)._cell_id

    def __lt__(self, other: CellID | CellIDLike) -> bool:
        return self._cell_tuple < CellID(other)._cell_tuple

    def __str__(self):
        return self._cell_id


class CellPointer:
    """Clase que representa un puntero a una celda en una hoja de cálculo
    Excel.

    Permite consultar o modificar el valor de la celda mediante 'value', y
    moverse por toda la hoja mediante los métodos 'up', 'down', 'left' y
    'right'. También permite moverse a una celda arbitraria mediante el método
    'goto'. Finalmente, se pueden crear copias de este puntero mediante el
    método 'copy'.

    No se recomienda crear instancias directamente de esta clase, sino usar
    el método 'pointer()' de la clase 'Excel'.

    """

    def __init__(self, sheet: OpenpyxlSheet, start_cell: CellID) -> None:
        self._sheet = sheet
        self.current = start_cell

    @property
    def cell(self) -> OpenpyxlCell:
        """Devuelve la celda actual, en formato openpyxl."""
        return self._sheet[self.current.as_str()]

    @property
    def value(self) -> Any:
        """Devuelve el valor de la celda actual."""
        return self.cell.value

    @value.setter
    def value(self, value: Any) -> None:
        """Modifica el valor de la celda actual."""
        self.cell.value = value

    # Movimientos
    def up(self, steps: int = 1) -> CellPointer:
        """Mueve el puntero hacia arriba."""
        self.current = CellID((self.current.as_tuple()[0], self.current.as_tuple()[1] - steps))
        return self

    def down(self, steps: int = 1) -> CellPointer:
        """Mueve el puntero hacia abajo."""
        self.current = CellID((self.current.as_tuple()[0], self.current.as_tuple()[1] + steps))
        return self

    def left(self, steps: int = 1) -> CellPointer:
        """Mueve el puntero hacia la izquierda."""
        self.current = CellID((self.current.as_tuple()[0] - steps, self.current.as_tuple()[1]))
        return self

    def right(self, steps: int = 1) -> CellPointer:
        """Mueve el puntero hacia la derecha."""
        self.current = CellID((self.current.as_tuple()[0] + steps, self.current.as_tuple()[1]))
        return self

    def goto(self, cell: CellIDLike | CellID) -> CellPointer:
        """Mueve el puntero a una celda arbitraria."""
        self.current = CellID(cell)
        return self

    # Otros
    def copy(self) -> CellPointer:
        """Crea una copia del puntero."""
        return CellPointer(self._sheet, self.current)

    def __str__(self):
        return f"CellPointer({self.current})"


class Excel:
    """Clase que representa un archivo de hoja de cálculo Excel.

    Esencialmente sirve para crear o abrir un archivo de hoja de cálculo Excel,
    y para crear punteros a celdas en una hoja de cálculo. También permite
    ejecutar operaciones de selección, creación y modificación de páginas; y
    también enmascara operaciones propias de openpyxl.

    El constructor requiere la ruta a un archivo de hoja de cálculo Excel. Si
    el archivo no existe, se creará uno nuevo. Si el archivo existe, se abrirá
    el archivo existente. Alternativamente, se pueden usar los métodos de clase
    'new' (que crea un archivo nuevo o sobrescribe uno existente) o 'open' (que
    abre un archivo existente o lanza una excepción si no existe).

    """

    def __init__(self, path: Path | str) -> None:
        self._path = Path(path)
        if self._path.exists():
            self._wb = openpyxl.load_workbook(self._path)
        else:
            self._wb = openpyxl.Workbook()
            self._wb.save(self._path)
        self._sheet = self._wb.active

    @classmethod
    def new(cls, path: Path | str) -> Excel:
        """Crea un archivo nuevo de hoja de cálculo Excel."""
        path = Path(path)
        if path.exists():
            path.unlink()
        return cls(path)

    @classmethod
    def open(cls, path: Path | str) -> Excel:
        """Abre un archivo existente de hoja de cálculo Excel."""
        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"No existe el archivo '{path}'.")
        return cls(path)

    def save(self) -> None:
        """Guarda el archivo de hoja de cálculo Excel."""
        self._wb.save(self._path)

    def close(self) -> None:
        """Cierra el archivo de hoja de cálculo Excel."""
        self._wb.close()

    # Métodos de páginas
    @property
    def current_sheet(self) -> OpenpyxlSheet:
        """Devuelve la página actual de la hoja de cálculo."""
        return self._sheet

    def select_sheet(self, id: int | str = 0) -> OpenpyxlSheet:
        """Cambia la página seleccionada.

        El identificador puede ser o bien el índice de la página (empezando por
        0), o bien el nombre de la página.

        """
        if isinstance(id, int):
            self._sheet = self._wb.worksheets[id]
        elif isinstance(id, str):
            self._sheet = self._wb[id]
        else:
            raise ValueError("El identificador de la página debe ser un entero o una cadena.")
        return self._sheet

    def new_sheet(self, name: str) -> OpenpyxlSheet:
        """Crea una nueva página en la hoja de cálculo.

        Le asigna el nombre especificado y se mueve a ella.

        """
        self._wb.create_sheet(name=name)
        self._sheet = self._wb[name]
        return self._sheet

    def rename_sheet(self, new_name: str) -> None:
        """Cambia el nombre de la página actualmente seleccionada."""
        self._sheet.title = new_name

    def delete_sheet(self) -> None:
        """Elimina la página actualmente seleccionada, a continuación, se mueve
        a la primera página.

        """
        to_del = self._sheet
        self.select_sheet(0)
        self._wb.remove(to_del)

    # Métodos de celdas
    def pointer(self, at: CellIDLike | CellID) -> CellPointer:
        """Crea un puntero a una celda en la página actual."""
        return CellPointer(self._sheet, CellID(at))

    # Utilidades y herramientas
    def compose_formula(self, formula: str, cells: list[CellIDLike | CellID]) -> str:
        """Compone una fórmula a partir de una lista de identificadores de
        celdas.

        La fórmula es la especificada en 'formula' y los identificadores de
        celdas son los especificados en 'cells'.

        La fórmula resultante es capaz de convertir la lista de identificadores
        en rangos de celdas allí donde sea posible.

        """
        cells = list(sorted(CellID(cell) for cell in cells))
        cell_ranges = []
        prev_cell = None
        for cell in cells:
            if prev_cell is None:
                prev_cell = cell
                cell_ranges.append([cell, None])
                continue
            x, y = cell.as_tuple()
            px, py = prev_cell.as_tuple()
            if abs(x - px) == 1 and y == py:
                cell_ranges[-1][1] = cell
            elif abs(y - py) == 1 and x == px:
                cell_ranges[-1][1] = cell
            else:
                cell_ranges.append([cell, None])
            prev_cell = cell
        formatted_cells = []
        for start, end in cell_ranges:
            if end is None:
                formatted_cells.append(start.as_str())
            else:
                formatted_cells.append(f"{start.as_str()}:{end.as_str()}")
        return f"={formula}({','.join(formatted_cells)})"

    # Otros
    def __str__(self):
        return f"Excel({self._path})"
