# 3.10.11
# Creado: 03/03/2024
"""Módulo de utilidades y herramientas para manejar hojas de cálculo de Excel
más cómodamente.

Por detrás usa la librería openpyxl para manejar los archivos de Excel. El
objetivo es que el manejo de una hoja de cálculo sea más fiel al que se haría
con una hoja de cálculo de Excel real.

"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import Cell as OpenpyxlCell
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlSheet

from marx.util.excel import CellStyle


CellIDLike = str | tuple[int, int]


# Hojas


class Sheet:
    """Clase que representa una página de una hoja de cálculo Excel."""

    def __init__(self, manager: ExcelManager, sheet: OpenpyxlSheet) -> None:
        self._manager = manager
        self.raw = sheet
        self.cells = {}

        # Configuración general
        self.raw.sheet_properties.outlinePr.summaryBelow = False

    @property
    def title(self) -> str:
        """Devuelve el nombre de la página."""
        return self.raw.title

    @title.setter
    def title(self, new_title: str) -> None:
        """Cambia el nombre de la página."""
        self.raw.title = new_title

    @property
    def index(self) -> int:
        """Devuelve el índice de la página en la hoja de cálculo."""
        return self._manager._wb.index(self.raw)

    # Selección y modificación
    def select(self) -> Sheet:
        """Selecciona la página actual."""
        self.raw.current_sheet = self
        return self

    def rename(self, new_title: str) -> Sheet:
        """Cambia el nombre de la página actual."""
        self.raw.title = new_title
        return self

    def clear(self) -> Sheet:
        """Elimina todas las celdas de la página actual."""
        self.raw.delete_rows(1, self.raw.max_row)
        return self

    def delete(self) -> None:
        """Elimina la página actual."""
        self._manager._wb.remove(self._sheet)

    # Configuración de página
    def set_column_width(self, column: str | int, width: int) -> None:
        """Establece el ancho de una columna."""
        if isinstance(column, int):
            column = chr(column + 64)
        self.raw.column_dimensions[column.upper()].width = width

    def group_rows(self, start: int, end: int, outline_level: int = 0) -> None:
        # TODO: Implementar
        """Agrupa un rango de filas."""
        print(f">>> [{outline_level}] {start} - {end}")
        self.raw.row_dimensions.group(start + 1, end, outline_level=outline_level)

    # Selección de celdas
    def __getitem__(self, key: CellIDLike | CellID) -> Cell:
        id = CellID(key)
        if id not in self.cells:
            self.cells[id] = Cell(id, self.raw[id.id])
        return self.cells[id]

    # Otros
    def __str__(self):
        return f"<Sheet '{self.title}' at {self.index}>"


class Sheets:
    """Clase que representa al conjunto de páginas de una hoja de cálculo."""

    def __init__(self, manager: ExcelManager) -> None:
        self._manager = manager
        self._sheets = []
        for sheetname in self._manager._wb.sheetnames:
            self._sheets.append(Sheet(self._manager, self._manager._wb[sheetname]))

    def new(self, title: str, position: int | None = None, *, select: bool = False) -> Sheet:
        """Crea una nueva página en la hoja de cálculo.

        Si se indica 'position', la nueva página se insertará en esa posición
        (empezando por 0). Si no se indica, la nueva página se insertará al
        final.

        Si 'select' es True, la nueva página se seleccionará tras su creación.

        Devuelve la nueva página.

        """
        self._manager._wb.create_sheet(title=title, index=position)
        new = Sheet(self._manager, self._manager._wb[title])
        self._sheets.insert(position or len(self), new)
        if select:
            new.select()
        return new

    def __getitem__(self, key: int | str) -> Sheet:
        """Escoge una página por su índice o su nombre."""
        if isinstance(key, str):
            key = self._manager._wb.sheetnames.index(key)
        return self._sheets[key]

    def __len__(self) -> int:
        return len(self._manager._wb.sheetnames)

    def __iter__(self):
        return iter(self._sheets)

    def __str__(self):
        return f"<Sheets of '{self._manager.path}' ({len(self)})>"


# Celdas


class CellID:
    """Clase que representa un identificador de una celda arbitraria en una
    hoja de cálculo Excel.

    Los identificadores de celda tienen dos formas: texto y tupla. Por ejemplo,
    la celda A1 puede ser representada como "A1" o (1, 1). Esta clase permite
    convertir entre ambas formas y también permite comparar entre ellas.

    El constructor de la clase acepta un identificador de celda en cualquiera
    de sus formas.

    """

    def __init__(self, id: CellIDLike | CellID) -> None:
        self._row = 0
        self._col = 0
        self._lcol = ""
        if isinstance(id, CellID):
            self._row = id.row
            self._col = id.column
            self._lcol = id.lcol
        elif isinstance(id, str):
            self._lcol = "".join(ch for ch in id if ch.isalpha())
            self._col = sum((ord(ch) - 64) * (26**i) for i, ch in enumerate(reversed(self._lcol)))
            self._row = int("".join(ch for ch in id if ch.isdigit()))
        elif isinstance(id, tuple):
            self._col = id[0]
            self._lcol = ""
            aux = id[0]
            while aux > 0:
                aux, rem = divmod(aux - 1, 26)
                self._lcol = f"{chr(rem + 65)}{self._lcol}"
            self._row = id[1]
        else:
            raise ValueError("El identificador de celda debe ser una cadena o una tupla.")

    # Identificador de celda
    @property
    def id(self) -> str:
        return f"{self._lcol}{self._row}"

    @property
    def coords(self) -> tuple[int, int]:
        return (self._col, self._row)

    # Coordenadas individuales
    @property
    def row(self) -> int:
        return self._row

    @property
    def column(self) -> int:
        return self._col

    @property
    def col(self) -> int:
        return self._col

    @property
    def lcol(self) -> str:
        """Letra de la columna"""
        return self._lcol

    # Comparación
    def __eq__(self, other: CellID | CellIDLike) -> bool:
        return self.id == CellID(other).id

    def __lt__(self, other: CellID | CellIDLike) -> bool:
        return self.coords < CellID(other).coords

    # Otros
    def __hash__(self) -> int:
        return hash(self.id)

    def __str__(self):
        return self.id


class Cell:
    """Clase que representa una celda en una hoja de cálculo Excel."""

    def __init__(self, id: CellID, cell: OpenpyxlCell, style: CellStyle | None = None) -> None:
        self.id = id
        self.raw = cell
        self.style = style or CellStyle()

    @property
    def value(self) -> Any:
        """Devuelve el valor de la celda."""
        return self.raw.value

    @value.setter
    def value(self, value: Any) -> None:
        """Modifica el valor de la celda."""
        self.raw.value = value

    def __str__(self):
        return f"<Cell {self.id} in {self.raw.parent.title!r}>"


class CellPointer:
    """Clase que representa un puntero a una celda en una hoja de cálculo
    Excel.

    Permite moverse por toda la hoja mediante los métodos 'up', 'down', 'left',
    'right' y 'goto'. Para acceder a la celda apuntada, se proporciona el
    atributo 'cell', de la clase 'Cell'.

    """

    def __init__(self, sheet: Sheet, start: CellID | CellIDLike = "A1") -> None:
        self.sheet = sheet
        self._pos = CellID(start)
        # Creamos la celda de partida si no existe
        self.sheet[self._pos]

    @property
    def current(self) -> CellID:
        """Devuelve el identificador de la celda apuntada por el puntero."""
        return self._pos

    @property
    def cell(self) -> Cell:
        """Devuelve la celda apuntada por el puntero."""
        return self.sheet.cells[self._pos]

    @property
    def value(self) -> Any:
        """Devuelve el valor de la celda apuntada por el puntero."""
        return self.cell.value

    @value.setter
    def value(self, value: Any) -> None:
        """Modifica el valor de la celda apuntada por el puntero."""
        self.cell.value = value

    @property
    def row(self) -> int:
        """Devuelve el número de fila de la celda apuntada por el puntero."""
        return self._pos.coords[1]

    @property
    def column(self) -> int:
        """Devuelve el número de columna de la celda apuntada por el puntero."""
        return self._pos.coords[0]

    # Movimientos
    def up(self, steps: int = 1) -> CellPointer:
        """Mueve el puntero hacia arriba."""
        return self.goto(CellID((self._pos.col, self._pos.row + steps)))

    def down(self, steps: int = 1) -> CellPointer:
        """Mueve el puntero hacia abajo."""
        return self.goto(CellID((self._pos.col, self._pos.row - steps)))

    def left(self, steps: int = 1) -> CellPointer:
        """Mueve el puntero hacia la izquierda."""
        return self.goto(CellID((self._pos.col - steps, self._pos.row)))

    def right(self, steps: int = 1) -> CellPointer:
        """Mueve el puntero hacia la derecha."""
        return self.goto(CellID((self._pos.col + steps, self._pos.row)))

    def goto(self, id: CellIDLike | CellID) -> CellPointer:
        """Mueve el puntero a una celda arbitraria."""
        self._pos = CellID(id)
        self.sheet[self._pos]  # Asegura que la celda esté en el diccionario
        return self

    # Otros
    def copy(self) -> CellPointer:
        """Crea una copia del puntero."""
        return CellPointer(self.sheet, self._pos)

    def __str__(self):
        return f"<CellPointer at {self._pos} in {self.sheet.title!r}>"


# Principal


class ExcelManager:
    """Clase que representa un archivo de hoja de cálculo Excel.

    Esencialmente sirve para crear o abrir un archivo de hoja de cálculo Excel,
    y para crear punteros a celdas en una hoja de cálculo. También permite
    ejecutar operaciones de selección, creación y modificación de páginas; y
    también enmascara operaciones propias de openpyxl.

    El constructor requiere la ruta a un archivo de hoja de cálculo Excel.
    Por defecto, si el archivo no existe, se creará uno nuevo. Si se indica
    'create' como False, lanzará un error si no existe.
    A su vez, por defecto, si el archivo existe, lo abrirá. Si se indica
    'overwrite' como True, lo sobreescribirá.
    Se puede indicar la página inicial con 'sheet', que puede ser o bien el
    nombre de la página, o bien el índice (empezando por 0).

    """

    def __init__(
        self,
        path: Path | str,
        *,
        sheet: str | int = 0,
        create: bool = True,
        overwrite: bool = False,
    ) -> None:
        # Carga el archivo de hoja de cálculo Excel
        self.path = Path(path)
        if self.path.exists():
            if not overwrite:
                self._wb = openpyxl.load_workbook(self.path)
            else:
                self._wb = openpyxl.Workbook()
        elif create:
            self._wb = openpyxl.Workbook()
        else:
            raise FileNotFoundError(f"No se encontró el archivo {self.path!r}.")

        # Inicializa los atributos necesarios
        self.sheets = Sheets(self)
        self.current_sheet = self.sheets[sheet]

    # Control general
    def save(self) -> None:
        """Guarda el archivo de hoja de cálculo Excel."""
        self._wb.save(self.path)

    def close(self) -> None:
        """Cierra el archivo de hoja de cálculo Excel."""
        self._wb.close()

    # Métodos de celdas
    def pointer(self, at: CellIDLike | CellID) -> CellPointer:
        """Crea un puntero a una celda en la página actual."""
        return CellPointer(self.current_sheet, CellID(at))

    # Estilos
    def stylize(self) -> None:
        """Aplica los estilos asignados a cada una de las celdas creadas."""
        for sheet in self.sheets:
            for cell in sheet.cells.values():
                cell.style.apply(cell.raw)

    # Otros
    def __str__(self):
        return f"<Excel file at {self.path}>"
